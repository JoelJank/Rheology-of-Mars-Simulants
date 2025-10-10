import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import savgol_filter
import math
import os 
import sys
import pandas as pd
from openpyxl import load_workbook
import json

def get_sheetnames(Filepath):
    wb = load_workbook(Filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def get_settings_file(Settings_path):
    with open(Settings_path,'r') as f:
        return json.load(f)

def get_excel_file(Excel_path, Sheet_name):
    file = pd.read_excel(Excel_path, sheet_name = Sheet_name,header = 0,names= ["No","Time","Gap","Normal Force","Normal Stress", "Torque", "Shear Stress","Rotational Speed"] , skiprows=[1,2])
    mask = file.apply(lambda row: any(isinstance(x, str) for x in row), axis=1)
    file = file[~mask].dropna().reset_index(drop=True)
    return file

plt.style.use('paper.mplstyle')

names = [
    [["Glass 3kPa", "Glass 6kPa", "Glass 9kPa", "Glass 15 kPa"], "$\phi = 0$"],
    [["Glass 15 kPa"], "$\phi = 0.05$"],
    [["Glass 3kPa", "Glass 6kPa", "Glass 9kPa"], "$\phi = 0.1$"],
    [["Glass 15kPa"], "$\phi = 0.15$"],
    [["Glass 3kPa", "Glass 6kPa", "Glass 9kPa", "Glass 15 kPa"], "$\phi = 0.2$"],
    [["Glass 15 kPa"], "$\phi = 0.25$"]
]

files = ["H:/AG Parteli/Paris/Rheology-of-Mars-Simulants/Analysis/raw_data/Excel/Glass_dried/Glass_dried_1runs_20250708.xlsx",
         "H:/AG Parteli/Paris/Rheology-of-Mars-Simulants/Analysis/raw_data/Excel/Glass_15kPa/Glass_5percentwater.xlsx",
         "H:/AG Parteli/Paris/Rheology-of-Mars-Simulants/Analysis/raw_data/Excel/Glass_10%_unmixed/Glass_10percentwater_unmixed_1runs_20250710.xlsx",
         "H:/AG Parteli/Paris/Rheology-of-Mars-Simulants/Analysis/raw_data/Excel/Glass_15kPa/Glass_15percentwater.xlsx",
         "H:/AG Parteli/Paris/Rheology-of-Mars-Simulants/Analysis/raw_data/Excel/Glass_20%_mixed/Glass_20percentwater_mixed_1runs_20250711.xlsx",
         "H:/AG Parteli/Paris/Rheology-of-Mars-Simulants/Analysis/raw_data/Excel/Glass_15kPa/Glass_25percentwater.xlsx"
         ]

fig, axes = plt.subplots(4,6, figsize=(25, 18))
print("Program started")
file_count = 0
for file in files:
    sheetnames = get_sheetnames(file)
    del sheetnames[:1]
    sheetcount = 0
    for sheet in sheetnames:
        
        df = get_excel_file(file, sheet)
        time_col = df["Time"].values
        jumps_where = np.where(np.abs(np.diff(time_col)) > 10)[0] #Determine switch between different consolidation forces
        jumps_end = jumps_where[-1]+jumps_where[0]+1
        jumps = np.append(jumps_where, jumps_end)
        jump_starts = np.array([0])
        for i in range(len(jumps)-1):
            jump_starts = np.append(jump_starts, jumps[i]+1)
        ax = axes[sheetcount,file_count]
        ax2 = ax.twinx()
        
        #Normal Stress
        p1, = ax.plot(df["Time"],df["Normal Stress"],marker="o",linestyle="None",markersize=1, color="blue",label="Normal Stress")
        for i in range(len(jump_starts)):
            mean = np.mean(df["Normal Stress"][jump_starts[i]:jumps[i]])
            ax.hlines(mean, xmin = df["Time"][jump_starts[i]], xmax = df["Time"][jumps[i]], color="green", linestyle="--")
        if file_count == 0:
            ax.set_ylabel("Normal Stress [Pa]"); ax.set_xlabel("Time [s]")
            ax.yaxis.label.set_color(p1.get_color())
            ax.tick_params(axis = 'y', colors = p1.get_color())
        ax.set_title(f"{names[file_count][0][sheetcount]}, {names[file_count][1]}")
        
        #Shear Stress with savgol
        p2, = ax2.plot(df["Time"],df["Shear Stress"],marker="o",linestyle = "None", markersize=1, color="red", label="Shear Stress")
        for i in range(len(jump_starts)):
            sav = savgol_filter(df["Shear Stress"][jump_starts[i]:jumps[i]], 5, 2)
            p2_sav, = ax2.plot(df["Time"][jump_starts[i]:jumps[i]], sav, color="purple", linestyle = "--")
            sav_max = [np.max(sav), list(sav).index(max(sav))]
            p2_max, = ax2.plot(df["Time"][jump_starts[i]+sav_max[1]], sav[sav_max[1]], marker="x", color = "black", linestyle = "None")
        if file_count == len(files)-1:
            ax2.set_ylabel("Shear Stress [Pa]")
            ax2.tick_params(axis = 'y', colors = p2.get_color())
            ax2.yaxis.label.set_color(p2.get_color())
        
        sheetcount += 1
    file_count += 1

plt.subplots_adjust(
    left=0.05,    # Linker Rand minimieren
    right=0.98,   # Rechter Rand minimieren  
    bottom=0.05,  # Unterer Rand minimieren
    top=0.95,     # Oberer Rand minimieren
    wspace=0.2,   # Horizontalen Abstand stark reduzieren
    hspace=0.15    # Vertikalen Abstand stark reduzieren
)
plt.savefig("Glass_all.svg", dpi=300)